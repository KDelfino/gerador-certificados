import os
import re
import csv
import openpyxl
import win32com.client
import pythoncom

def limpar_nome_arquivo(nome):
    """
    Remove caracteres inválidos para nomes de arquivos no Windows.
    """
    if not nome:
        return ""
    # Remove caracteres inválidos
    nome_limpo = re.sub(r'[\\/*?:"<>|]', "", str(nome))
    # Limpa espaços em branco extras
    return nome_limpo.strip()

def ler_colunas_planilha(caminho):
    """
    Retorna a lista de cabeçalhos da planilha (Excel ou CSV).
    """
    if not os.path.exists(caminho):
        return []
        
    caminho_lower = caminho.lower()
    if caminho_lower.endswith('.csv'):
        try:
            with open(caminho, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                headers = next(reader)
                return [h.strip() for h in headers if h.strip()]
        except UnicodeDecodeError:
            # Tenta ler com encoding latin-1 se falhar utf-8
            with open(caminho, 'r', encoding='latin-1') as f:
                reader = csv.reader(f)
                headers = next(reader)
                return [h.strip() for h in headers if h.strip()]
        except Exception:
            return []
    else:
        try:
            wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
            sheet = wb.active
            for row in sheet.iter_rows(max_row=1, values_only=True):
                return [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
        except Exception:
            return []
    return []

def ler_dados_planilha(caminho):
    """
    Lê a planilha (Excel ou CSV) e retorna uma lista de dicionários.
    """
    if not os.path.exists(caminho):
        return []

    caminho_lower = caminho.lower()
    rows = []
    
    if caminho_lower.endswith('.csv'):
        encodings = ['utf-8', 'latin-1', 'cp1252']
        reader = None
        for enc in encodings:
            try:
                # Primeiro lê para verificar
                with open(caminho, 'r', encoding=enc) as f:
                    # Detecta delimitador (, ou ;)
                    sample = f.read(2048)
                    f.seek(0)
                    delimiter = ';' if ';' in sample else ','
                    
                    reader = csv.DictReader(f, delimiter=delimiter)
                    for row in reader:
                        cleaned_row = {str(k).strip(): str(v).strip() for k, v in row.items() if k is not None}
                        rows.append(cleaned_row)
                break
            except (UnicodeDecodeError, csv.Error):
                rows = []
                continue
    else:
        try:
            wb = openpyxl.load_workbook(caminho, data_only=True)
            sheet = wb.active
            headers = []
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(cell).strip() if cell is not None else f"Coluna_{idx}" for idx, cell in enumerate(row)]
                else:
                    if any(cell is not None for cell in row):
                        row_dict = {}
                        for idx, cell in enumerate(row):
                            if idx < len(headers):
                                val = str(cell).strip() if cell is not None else ""
                                row_dict[headers[idx]] = val
                        rows.append(row_dict)
        except Exception as e:
            print(f"Erro ao ler Excel: {e}")
            return []
            
    return rows

def obter_quantidade_slides(caminho_pptx):
    """
    Retorna a quantidade de slides em uma apresentação PowerPoint (.pptx).
    """
    pythoncom.CoInitialize()
    powerpoint = None
    presentation = None
    count = 0
    try:
        powerpoint = win32com.client.Dispatch("PowerPoint.Application")
        # WithWindow=False roda em background
        presentation = powerpoint.Presentations.Open(caminho_pptx, ReadOnly=True, WithWindow=False)
        count = presentation.Slides.Count
    except Exception as e:
        print(f"Erro ao contar slides: {e}")
        raise e
    finally:
        if presentation:
            try:
                presentation.Close()
            except:
                pass
        if powerpoint:
            try:
                powerpoint.Quit()
            except:
                pass
        pythoncom.CoUninitialize()
    return count

def escanear_placeholders_slide(caminho_pptx, slide_index):
    """
    Retorna uma lista de placeholders (#texto) encontrados no slide específico.
    """
    pythoncom.CoInitialize()
    powerpoint = None
    presentation = None
    placeholders = set()
    
    try:
        powerpoint = win32com.client.Dispatch("PowerPoint.Application")
        presentation = powerpoint.Presentations.Open(caminho_pptx, ReadOnly=True, WithWindow=False)
        
        if slide_index < 1 or slide_index > presentation.Slides.Count:
            raise ValueError(f"Índice de slide inválido: {slide_index}")
            
        slide = presentation.Slides(slide_index)
        
        def scan_shapes(shapes):
            for shape in shapes:
                # msoGroup = 6
                if shape.Type == 6:
                    scan_shapes(shape.GroupItems)
                # msoTable = 19
                elif shape.HasTable:
                    table = shape.Table
                    for row in table.Rows:
                        for cell in row.Cells:
                            if cell.Shape.HasTextFrame and cell.Shape.TextFrame.HasText:
                                text = cell.Shape.TextFrame.TextRange.Text
                                found = re.findall(r'#[\w_]+', text)
                                placeholders.update(found)
                elif shape.HasTextFrame:
                    if shape.TextFrame.HasText:
                        text = shape.TextFrame.TextRange.Text
                        found = re.findall(r'#[\w_]+', text)
                        placeholders.update(found)
                        
        scan_shapes(slide.Shapes)
        print(f"[DIAGNÓSTICO] escanear_placeholders_slide chamado!")
        print(f"[DIAGNÓSTICO] Caminho: {caminho_pptx}")
        print(f"[DIAGNÓSTICO] Slide Index: {slide_index}")
        print(f"[DIAGNÓSTICO] Slides Totais no arquivo: {presentation.Slides.Count}")
        print(f"[DIAGNÓSTICO] Encontrados: {sorted(list(placeholders))}")
    except Exception as e:
        print(f"Erro ao escanear placeholders: {e}")
        raise e
    finally:
        if presentation:
            try:
                presentation.Close()
            except:
                pass
        if powerpoint:
            try:
                powerpoint.Quit()
            except:
                pass
        pythoncom.CoUninitialize()
        
    return sorted(list(placeholders))

def substituir_texto_shapes(shapes, texto_antigo, texto_novo):
    """
    Substitui recursivamente texto_antigo por texto_novo preservando a formatação.
    """
    for shape in shapes:
        # Grupo de formas (msoGroup = 6)
        if shape.Type == 6:
            substituir_texto_shapes(shape.GroupItems, texto_antigo, texto_novo)
        # Tabela (msoTable = 19)
        elif shape.HasTable:
            table = shape.Table
            for row in table.Rows:
                for cell in row.Cells:
                    if cell.Shape.HasTextFrame and cell.Shape.TextFrame.HasText:
                        text_range = cell.Shape.TextFrame.TextRange
                        if texto_antigo in text_range.Text:
                            text_range.Replace(texto_antigo, texto_novo)
        # Caixa de texto normal
        elif shape.HasTextFrame:
            if shape.TextFrame.HasText:
                text_range = shape.TextFrame.TextRange
                if texto_antigo in text_range.Text:
                    text_range.Replace(texto_antigo, texto_novo)

def gerar_certificados_job(
    caminho_pptx,
    slide_index,
    dados_planilha,
    mapeamento,
    formato_saida,
    pasta_destino,
    padrao_nome,
    progress_callback=None,
    is_cancelled_callback=None
):
    """
    Job principal executado na Thread para gerar certificados em lote.
    """
    pythoncom.CoInitialize()
    
    # Inicia a automação COM
    powerpoint = win32com.client.Dispatch("PowerPoint.Application")
    powerpoint.Visible = True # Algumas chamadas COM exigem visibilidade no Windows
    
    try:
        total = len(dados_planilha)
        for idx, linha in enumerate(dados_planilha, 1):
            if is_cancelled_callback and is_cancelled_callback():
                if progress_callback:
                    progress_callback(idx - 1, total, "Geração cancelada pelo usuário.")
                break
                
            # Abre o modelo de forma limpa a cada iteração
            presentation = powerpoint.Presentations.Open(caminho_pptx, ReadOnly=True, WithWindow=False)
            
            # Se a saída for PDF, removemos todos os outros slides para ter um PDF com apenas o slide desejado
            if formato_saida == 'PDF':
                for j in range(presentation.Slides.Count, 0, -1):
                    if j != slide_index:
                        presentation.Slides(j).Delete()
                slide = presentation.Slides(1)
            else:
                slide = presentation.Slides(slide_index)
                
            # Faz as substituições com base no mapeamento
            for placeholder, coluna in mapeamento.items():
                valor = str(linha.get(coluna, ''))
                substituir_texto_shapes(slide.Shapes, placeholder, valor)
                
            # Resolve o nome do arquivo final
            nome_final = padrao_nome
            for coluna, valor in linha.items():
                marcador = "{" + coluna + "}"
                if marcador in nome_final:
                    nome_final = nome_final.replace(marcador, str(valor))
                    
            # Limpa caracteres inválidos
            nome_final = limpar_nome_arquivo(nome_final)
            if not nome_final:
                nome_final = f"Certificado_{idx}"
                
            extensao = ".pdf" if formato_saida == 'PDF' else ".png"
            caminho_saida = os.path.join(pasta_destino, nome_final + extensao)
            
            # Exporta o slide correspondente
            if formato_saida == 'PDF':
                # PpSaveAsFileType.ppSaveAsPDF = 32
                presentation.SaveAs(caminho_saida, 32)
            else:
                # Exporta slide como PNG
                slide.Export(caminho_saida, "PNG")
                
            # Fecha a apresentação descartando mudanças
            presentation.Saved = True
            presentation.Close()
            
            if progress_callback:
                progress_callback(idx, total, f"Gerado [{idx}/{total}]: {nome_final}{extensao}")
                
    except Exception as e:
        print(f"Erro no loop de geração: {e}")
        raise e
    finally:
        if powerpoint:
            try:
                powerpoint.Quit()
            except:
                pass
        pythoncom.CoUninitialize()
